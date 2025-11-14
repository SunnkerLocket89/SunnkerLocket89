"""Command line helper for processing the Idaho4 exhibits spreadsheet.

The parser performs the following steps:

* Load an Excel workbook that contains the Idaho4 exhibit metadata.
* Determine which column contains the PDF download URL and which column
  uniquely identifies the exhibit.  Both columns can be provided via
  command line flags and are automatically inferred when omitted.
* Download the PDF assets referenced in the spreadsheet.
* Optionally extract the first N pages of each PDF into a separate file.
* Persist a JSON manifest that captures the processing result for each
  exhibit row.

The tool is intentionally defensive: every error is captured and stored in
the manifest so the caller can inspect what went wrong without losing
successful downloads.  The script is designed to be re-entrant which allows
the caller to resume interrupted downloads without re-fetching previously
processed files.
"""

from __future__ import annotations

import argparse
import concurrent.futures
import csv
import json
import logging
import math
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Sequence, Tuple
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from xml.etree import ElementTree
from zipfile import ZipFile

try:  # ``requests`` handles the HTTP downloads when available.
    import requests
    from requests import Response
except Exception:  # pragma: no cover - optional dependency
    requests = None  # type: ignore[assignment]
    Response = object  # type: ignore[assignment]

try:  # ``tqdm`` is used to render progress bars while processing rows.
    from tqdm import tqdm
except Exception:  # pragma: no cover - optional dependency

    def tqdm(iterable: Iterable[Any], *args: Any, **kwargs: Any) -> Iterator[Any]:
        """Lightweight stand-in for :func:`tqdm.tqdm` when the package is missing."""

        total = kwargs.get("total")
        desc = kwargs.get("desc", "")
        count = 0
        for item in iterable:
            count += 1
            if total:
                message = f"\r{desc}: {count}/{total}"
                sys.stderr.write(message)
                sys.stderr.flush()
            yield item
        if total:
            sys.stderr.write("\n")
        return


try:  # ``openpyxl`` is used to read the input workbook when installed.
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None  # type: ignore[assignment]

try:  # PyMuPDF is the preferred backend for page extraction.
    import fitz  # type: ignore[attr-defined]
except ImportError:  # pragma: no cover - optional dependency
    fitz = None  # type: ignore[assignment]

try:  # PyPDF2 is used as a fallback if PyMuPDF is unavailable.
    from PyPDF2 import PdfReader, PdfWriter  # type: ignore
except ImportError:  # pragma: no cover - optional dependency
    PdfReader = None  # type: ignore[assignment]
    PdfWriter = None  # type: ignore[assignment]


LOGGER = logging.getLogger(__name__)


def configure_logging(verbose: bool) -> None:
    """Configure the global logging level."""

    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="[%(levelname)s] %(message)s",
    )


def infer_column(columns: Sequence[str], keywords: Sequence[str]) -> Optional[str]:
    """Return the first column name containing any of *keywords*.

    Columns are compared in a case-insensitive manner.  The helper allows the
    command line interface to provide sensible defaults when the user does not
    specify the relevant column names explicitly.
    """

    for keyword in keywords:
        for column in columns:
            if keyword in column.lower():
                return column
    return None


def safe_filename(value: str, *, fallback: str) -> str:
    """Return a filesystem-friendly representation of *value*.

    Non alphanumeric characters are replaced with underscores and the result
    is trimmed to a sensible length.  The function never returns an empty
    string and instead falls back to the *fallback* parameter.
    """

    value = value.strip()
    if not value:
        value = fallback

    sanitized = re.sub(r"[^A-Za-z0-9_.-]+", "_", value)
    sanitized = sanitized.strip("._")
    if not sanitized:
        sanitized = fallback
    return sanitized[:150]


def build_output_stem(identifier: str, row_number: int) -> str:
    """Return a stable filename stem for *identifier* in *row_number*.

    Including the row number guarantees uniqueness even when multiple
    spreadsheet rows share the same identifier or when different values are
    normalised to an identical filename.  The zero-padded representation keeps
    the files sorted in the same order as the input worksheet.
    """

    prefix = f"{row_number:05d}"
    return f"{prefix}_{identifier}"[:180]


def coerce_to_optional_str(value: Any) -> Optional[str]:
    """Convert *value* into a trimmed string or ``None`` when empty."""

    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return str(value).strip() or None


def clean_metadata(record: Dict[str, Any]) -> Dict[str, Any]:
    """Return a copy of *record* with NaN values converted to ``None``."""

    cleaned: Dict[str, Any] = {}
    for key, value in record.items():
        if isinstance(value, float) and math.isnan(value):
            cleaned[key] = None
        else:
            cleaned[key] = value
    return cleaned


@dataclass(slots=True)
class ProcessorConfig:
    """Configuration required for exhibit processing."""

    url_column: str
    id_column: str
    out_dir: Path
    downloads_dir: Path
    extracts_dir: Path
    extract_pages: int
    resume: bool
    timeout: int
    chunk_size: int = 1024 * 128


@dataclass(slots=True)
class ProcessResult:
    """Container describing the outcome for a single spreadsheet row."""

    row_number: int
    identifier: str
    status: str
    message: str
    metadata: Dict[str, Any]
    pdf_path: Optional[str]
    extracted_path: Optional[str]


@dataclass(slots=True)
class WorkbookData:
    """Container for the parsed Excel worksheet."""

    columns: List[str]
    records: List[Dict[str, Any]]
    row_numbers: List[int]


def _normalise_column_name(value: Any, index: int) -> str:
    """Return a safe column name derived from *value* or a fallback."""

    if value is None:
        return f"column_{index + 1}"
    if isinstance(value, str):
        name = value.strip()
        return name or f"column_{index + 1}"
    return str(value).strip() or f"column_{index + 1}"


def _strip_namespace(tag: str) -> str:
    """Remove the XML namespace from *tag* if present."""

    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def _column_index_from_ref(ref: str) -> int:
    """Return the zero-based column index extracted from an Excel cell ref."""

    letters = []
    for char in ref:
        if char.isalpha():
            letters.append(char.upper())
        else:
            break
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def _read_shared_strings(zf: ZipFile) -> List[str]:
    """Return the workbook shared strings table."""

    try:
        data = zf.read("xl/sharedStrings.xml")
    except KeyError:
        return []

    root = ElementTree.fromstring(data)
    strings: List[str] = []
    for element in root:
        if _strip_namespace(element.tag) != "si":
            continue
        fragments: List[str] = []
        for text_node in element.iter():
            if _strip_namespace(text_node.tag) == "t" and text_node.text is not None:
                fragments.append(text_node.text)
        strings.append("".join(fragments))
    return strings


def _resolve_sheet_targets(zf: ZipFile) -> List[Tuple[str, str]]:
    """Return a mapping of sheet name to XML path within the archive."""

    try:
        workbook_data = zf.read("xl/workbook.xml")
        rels_data = zf.read("xl/_rels/workbook.xml.rels")
    except KeyError as error:
        raise ValueError("The workbook is missing required metadata") from error

    root = ElementTree.fromstring(workbook_data)
    namespace = root.tag.split("}")[0].strip("{") if "}" in root.tag else ""
    sheets: List[Tuple[str, str]] = []
    rel_namespace = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    rels_root = ElementTree.fromstring(rels_data)
    relationships = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root
        if _strip_namespace(rel.tag) == "Relationship"
    }

    sheet_tag = f"{{{namespace}}}sheet" if namespace else "sheet"
    rel_key = f"{{{rel_namespace}}}id"
    for sheet in root.findall(f".//{sheet_tag}"):
        name = sheet.attrib.get("name", "")
        rel_id = sheet.attrib.get(rel_key)
        if not rel_id:
            continue
        target = relationships.get(rel_id)
        if not target:
            continue
        if target.startswith("../"):
            target = target.replace("../", "", 1)
        if target.startswith("/"):
            target = target.lstrip("/")
        elif not target.startswith("xl/"):
            target = f"xl/{target}"
        sheets.append((name, target))
    return sheets


def _parse_cell_value(cell: ElementTree.Element, shared_strings: Sequence[str], namespace: str) -> Any:
    """Return the Python value stored in *cell*."""

    cell_type = cell.attrib.get("t")
    value_element = cell.find(f"{{{namespace}}}v") if namespace else cell.find("v")

    if cell_type == "s":
        if value_element is None or value_element.text is None:
            return None
        try:
            index = int(value_element.text)
        except ValueError:
            return None
        return shared_strings[index] if 0 <= index < len(shared_strings) else None

    if cell_type == "inlineStr":
        inline = cell.find(f"{{{namespace}}}is") if namespace else cell.find("is")
        if inline is None:
            inline = cell
        fragments: List[str] = []
        for node in inline.iter():
            if _strip_namespace(node.tag) == "t" and node.text is not None:
                fragments.append(node.text)
        return "".join(fragments)

    if cell_type == "b":
        if value_element is None or value_element.text is None:
            return None
        return value_element.text == "1"

    if cell_type == "str":
        if value_element is None:
            return ""
        return value_element.text or ""

    if value_element is None or value_element.text is None:
        return None

    text = value_element.text.strip()
    if not text:
        return None

    try:
        if any(character in text for character in (".", "e", "E")):
            number = float(text)
            if number.is_integer():
                return int(number)
            return number
        return int(text)
    except ValueError:
        return text


def _iter_xlsx_rows(zf: ZipFile, sheet_path: str, shared_strings: Sequence[str]) -> Iterator[Tuple[int, Dict[int, Any]]]:
    """Yield row number and cell mapping for *sheet_path* inside *zf*."""

    try:
        data = zf.read(sheet_path)
    except KeyError as error:
        raise ValueError(f"Worksheet data '{sheet_path}' is missing from the workbook") from error
    root = ElementTree.fromstring(data)
    namespace = root.tag.split("}")[0].strip("{") if "}" in root.tag else ""
    sheet_data = root.find(f"{{{namespace}}}sheetData") if namespace else root.find("sheetData")
    if sheet_data is None:
        return iter(())

    def row_iterator() -> Iterator[Tuple[int, Dict[int, Any]]]:
        for row in sheet_data:
            if _strip_namespace(row.tag) != "row":
                continue
            try:
                row_number = int(row.attrib.get("r", "0"))
            except ValueError:
                row_number = 0
            last_index = -1
            cells: Dict[int, Any] = {}
            for cell in row:
                if _strip_namespace(cell.tag) != "c":
                    continue
                ref = cell.attrib.get("r")
                if ref:
                    column_index = _column_index_from_ref(ref)
                else:
                    column_index = last_index + 1
                last_index = column_index
                cells[column_index] = _parse_cell_value(cell, shared_strings, namespace)
            yield row_number or 0, cells

    return row_iterator()


def _read_workbook_via_builtin(path: Path, sheet: Optional[str]) -> WorkbookData:
    """Fallback Excel reader that only relies on the Python standard library."""

    with ZipFile(path) as zf:
        sheet_targets = _resolve_sheet_targets(zf)
        if not sheet_targets:
            raise ValueError("The workbook does not contain any worksheets")

        if sheet is None:
            _, sheet_path = sheet_targets[0]
        else:
            sheet_path = None
            for name, candidate in sheet_targets:
                if name == sheet or name.lower() == sheet.lower():
                    sheet_path = candidate
                    break
            if sheet_path is None:
                raise ValueError(f"Worksheet '{sheet}' not found in workbook")

        shared_strings = _read_shared_strings(zf)
        rows = list(_iter_xlsx_rows(zf, sheet_path, shared_strings))

    if not rows:
        raise ValueError("The worksheet does not contain any rows")

    header_row_number, header_cells = rows[0]
    if not header_cells:
        raise ValueError("The worksheet does not contain a header row")

    max_header_index = max(header_cells)
    columns = [
        _normalise_column_name(header_cells.get(index), index)
        for index in range(max_header_index + 1)
    ]

    records: List[Dict[str, Any]] = []
    row_numbers: List[int] = []
    for row_number, cells in rows[1:]:
        if cells:
            max_index = max(cells)
            if max_index >= len(columns):
                for index in range(len(columns), max_index + 1):
                    column_name = _normalise_column_name(None, index)
                    columns.append(column_name)
                    for record in records:
                        record[column_name] = None
        values = [cells.get(index) for index in range(len(columns))]
        if all(value is None for value in values):
            continue
        record = {columns[index]: values[index] for index in range(len(columns))}
        records.append(record)
        if row_number:
            resolved_row_number = row_number
        elif row_numbers:
            resolved_row_number = row_numbers[-1] + 1
        else:
            resolved_row_number = (header_row_number or 1) + len(records)
        row_numbers.append(resolved_row_number)

    return WorkbookData(columns=columns, records=records, row_numbers=row_numbers)


def read_input_workbook(path: Path, sheet: Optional[str]) -> WorkbookData:
    """Load *path* and return its header columns and row dictionaries."""

    if not path.exists():
        raise FileNotFoundError(f"Input workbook not found: {path}")

    LOGGER.debug("Loading workbook %s", path)

    if load_workbook is None:
        return _read_workbook_via_builtin(path, sheet)

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet] if sheet else workbook.active
    except KeyError as error:
        workbook.close()
        raise ValueError(f"Worksheet '{sheet}' not found in workbook") from error

    rows = worksheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration as error:
        workbook.close()
        raise ValueError("The worksheet does not contain any rows") from error

    columns = [_normalise_column_name(value, index) for index, value in enumerate(header)]

    records: List[Dict[str, Any]] = []
    row_numbers: List[int] = []
    for row_number, row in enumerate(rows, start=2):
        if row is None:
            continue
        values = list(row)
        if all(value is None for value in values):
            continue
        record: Dict[str, Any] = {}
        for index, column in enumerate(columns):
            record[column] = values[index] if index < len(values) else None
        records.append(record)
        row_numbers.append(row_number)

    workbook.close()

    return WorkbookData(columns=columns, records=records, row_numbers=row_numbers)


def download_pdf(url: str, destination: Path, *, resume: bool, timeout: int, chunk_size: int) -> bool:
    """Download *url* to *destination*.

    The download is skipped when *resume* is true and the destination already
    exists.  The function returns ``True`` when a fresh download occurred and
    ``False`` when the file was already present.
    """

    if destination.exists():
        if resume:
            LOGGER.debug("Skipping download for %s (already exists)", destination)
            return False
        destination.unlink()

    LOGGER.debug("Downloading %s -> %s", url, destination)

    if requests is not None:
        response: Response = requests.get(url, stream=True, timeout=timeout)
        response.raise_for_status()

        with destination.open("wb") as handle:
            for chunk in response.iter_content(chunk_size=chunk_size):
                if chunk:
                    handle.write(chunk)
        return True

    request = Request(url, headers={"User-Agent": "Python-urllib/3"})
    try:
        with urlopen(request, timeout=timeout) as source, destination.open("wb") as handle:
            while True:
                chunk = source.read(chunk_size)
                if not chunk:
                    break
                handle.write(chunk)
    except HTTPError as error:  # pragma: no cover - network interaction
        raise RuntimeError(f"HTTP error {error.code}: {error.reason}") from error
    except URLError as error:  # pragma: no cover - network interaction
        raise RuntimeError(f"Failed to download file: {error.reason}") from error

    return True


def extract_first_pages(pdf_path: Path, destination: Path, pages: int) -> None:
    """Extract the first *pages* pages from *pdf_path* into *destination*.

    PyMuPDF is used when available because it is fast and feature complete.
    PyPDF2 is used as a fallback.
    """

    if pages <= 0:
        return

    destination.parent.mkdir(parents=True, exist_ok=True)

    if fitz is not None:
        LOGGER.debug("Extracting %s page(s) from %s using PyMuPDF", pages, pdf_path)
        with fitz.open(pdf_path) as source:
            pages_to_copy = min(pages, source.page_count)
            with fitz.open() as target:
                for page_index in range(pages_to_copy):
                    target.insert_pdf(source, from_page=page_index, to_page=page_index)
                target.save(destination)
        return

    if PdfReader is None or PdfWriter is None:
        raise RuntimeError(
            "No PDF extraction backend is available. Install PyMuPDF or PyPDF2."
        )

    LOGGER.debug("Extracting %s page(s) from %s using PyPDF2", pages, pdf_path)
    reader = PdfReader(str(pdf_path))
    writer = PdfWriter()
    pages_to_copy = min(pages, len(reader.pages))
    for page_index in range(pages_to_copy):
        writer.add_page(reader.pages[page_index])
    with destination.open("wb") as handle:
        writer.write(handle)


def process_record(
    row_number: int,
    record: Dict[str, Any],
    config: ProcessorConfig,
) -> ProcessResult:
    """Download and optionally extract the exhibit described by *record*."""

    metadata = clean_metadata(record)

    identifier_value = coerce_to_optional_str(metadata.get(config.id_column))
    identifier = safe_filename(
        identifier_value or f"row-{row_number}",
        fallback=f"row-{row_number}",
    )

    output_stem = build_output_stem(identifier, row_number)

    url_value = coerce_to_optional_str(metadata.get(config.url_column))
    if not url_value:
        return ProcessResult(
            row_number=row_number,
            identifier=identifier,
            status="skipped",
            message=f"Missing URL in column '{config.url_column}'",
            metadata=metadata,
            pdf_path=None,
            extracted_path=None,
        )

    downloads_dir = config.downloads_dir
    downloads_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = downloads_dir / f"{output_stem}.pdf"

    try:
        fresh_download = download_pdf(
            url_value,
            pdf_path,
            resume=config.resume,
            timeout=config.timeout,
            chunk_size=config.chunk_size,
        )
    except Exception as error:  # pragma: no cover - network interaction
        LOGGER.debug("Download failed for %s: %s", identifier, error)
        return ProcessResult(
            row_number=row_number,
            identifier=identifier,
            status="error",
            message=f"Failed to download PDF: {error}",
            metadata=metadata,
            pdf_path=None,
            extracted_path=None,
        )

    extracted_path: Optional[Path] = None
    if config.extract_pages > 0:
        extracted_path = (
            config.extracts_dir
            / f"{output_stem}_first_{config.extract_pages}_pages.pdf"
        )
        try:
            extract_first_pages(pdf_path, extracted_path, config.extract_pages)
        except Exception as error:  # pragma: no cover - PDF backend interaction
            LOGGER.debug("Extraction failed for %s: %s", identifier, error)
            return ProcessResult(
                row_number=row_number,
                identifier=identifier,
                status="error",
                message=f"Failed to extract pages: {error}",
                metadata=metadata,
                pdf_path=str(pdf_path),
                extracted_path=None,
            )

    message = "Downloaded" if fresh_download else "Existing file reused"
    if config.extract_pages > 0:
        message = f"{message}; extracted first {config.extract_pages} page(s)"

    return ProcessResult(
        row_number=row_number,
        identifier=identifier,
        status="success",
        message=message,
        metadata=metadata,
        pdf_path=str(pdf_path),
        extracted_path=str(extracted_path) if extracted_path else None,
    )


def build_manifest(results: Sequence[ProcessResult]) -> Dict[str, Any]:
    """Create a JSON serialisable manifest from *results*."""

    entries: List[Dict[str, Any]] = []
    for result in results:
        entry = {
            "row_number": result.row_number,
            "identifier": result.identifier,
            "status": result.status,
            "message": result.message,
            "pdf_path": result.pdf_path,
            "extracted_path": result.extracted_path,
            "metadata": result.metadata,
        }
        entries.append(entry)

    totals = {
        "success": sum(1 for entry in entries if entry["status"] == "success"),
        "skipped": sum(1 for entry in entries if entry["status"] == "skipped"),
        "error": sum(1 for entry in entries if entry["status"] == "error"),
        "total": len(entries),
    }

    return {
        "summary": totals,
        "results": entries,
    }


def dump_manifest(manifest: Dict[str, Any], destination: Path) -> None:
    """Persist *manifest* as JSON at *destination*."""

    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8") as handle:
        json.dump(manifest, handle, indent=2, ensure_ascii=False)
        handle.write("\n")


def write_csv_summary(manifest: Dict[str, Any], destination: Path) -> None:
    """Write a CSV summary of *manifest* to *destination*."""

    destination.parent.mkdir(parents=True, exist_ok=True)

    base_fields = [
        "row_number",
        "identifier",
        "status",
        "message",
        "pdf_path",
        "extracted_path",
    ]

    metadata_keys: List[str] = []
    seen_keys: set[str] = set()
    for entry in manifest["results"]:
        metadata = entry.get("metadata", {}) or {}
        for key in metadata.keys():
            if key not in seen_keys and key not in base_fields:
                seen_keys.add(key)
                metadata_keys.append(key)

    fieldnames = base_fields + metadata_keys

    with destination.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for entry in manifest["results"]:
            row = {name: None for name in fieldnames}
            row.update({
                "row_number": entry["row_number"],
                "identifier": entry["identifier"],
                "status": entry["status"],
                "message": entry["message"],
                "pdf_path": entry["pdf_path"],
                "extracted_path": entry["extracted_path"],
            })
            metadata = entry.get("metadata", {}) or {}
            for key in metadata_keys:
                if key in metadata:
                    row[key] = metadata[key]
            writer.writerow(row)


def parse_arguments(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    """Return the parsed command line arguments."""

    parser = argparse.ArgumentParser(
        description="Download and extract PDF exhibits listed in an Excel workbook",
    )
    parser.add_argument(
        "--in-file",
        required=True,
        type=Path,
        help="Path to the Idaho4 exhibits Excel workbook.",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Worksheet name inside the workbook (defaults to the first sheet).",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=Path("idaho4_output"),
        help="Directory where downloads, extracts, and manifests are stored.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=4,
        help="Number of worker threads to use for concurrent downloads.",
    )
    parser.add_argument(
        "--extract-pages",
        type=int,
        default=0,
        help="Number of pages to extract from each PDF (0 disables extraction).",
    )
    parser.add_argument(
        "--url-column",
        type=str,
        default=None,
        help="Column containing the PDF URL (auto-detected when omitted).",
    )
    parser.add_argument(
        "--id-column",
        type=str,
        default=None,
        help="Column containing the exhibit identifier (auto-detected when omitted).",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=60,
        help="HTTP timeout in seconds for downloading each PDF.",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Skip downloading files that already exist in the output directory.",
    )
    parser.add_argument(
        "--manifest",
        type=Path,
        default=None,
        help="Optional path for the JSON manifest (defaults to out-dir/manifest.json).",
    )
    parser.add_argument(
        "--csv",
        type=Path,
        default=None,
        help="Optional path for a CSV summary (defaults to out-dir/manifest.csv).",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable verbose logging for troubleshooting.",
    )

    return parser.parse_args(argv)


def prepare_configuration(args: argparse.Namespace, workbook: WorkbookData) -> ProcessorConfig:
    """Create a :class:`ProcessorConfig` from the parsed *args*."""

    columns = [str(column) for column in workbook.columns]

    url_column = args.url_column or infer_column(columns, ("pdf url", "url", "link"))
    if url_column is None:
        raise ValueError(
            "Unable to infer the URL column. Use --url-column to specify it explicitly."
        )

    id_column = args.id_column or infer_column(columns, ("exhibit", "id", "identifier", "name", "title"))
    if id_column is None:
        # Fallback to the URL column if no identifier-like column exists.
        id_column = url_column

    out_dir = args.out_dir.resolve()
    downloads_dir = out_dir / "downloads"
    extracts_dir = out_dir / "extracted_pages"

    if args.extract_pages > 0:
        extracts_dir.mkdir(parents=True, exist_ok=True)

    return ProcessorConfig(
        url_column=url_column,
        id_column=id_column,
        out_dir=out_dir,
        downloads_dir=downloads_dir,
        extracts_dir=extracts_dir,
        extract_pages=max(0, int(args.extract_pages)),
        resume=bool(args.resume),
        timeout=max(1, int(args.timeout)),
    )


def run(argv: Optional[Sequence[str]] = None) -> int:
    """Entry point used by the ``__main__`` guard."""

    args = parse_arguments(argv)
    configure_logging(args.verbose)

    try:
        workbook = read_input_workbook(args.in_file, args.sheet)
    except Exception as error:
        LOGGER.error("Failed to load workbook: %s", error)
        return 1

    try:
        config = prepare_configuration(args, workbook)
    except Exception as error:
        LOGGER.error("%s", error)
        return 1

    records = list(workbook.records)
    row_numbers = list(workbook.row_numbers)
    total_records = len(records)
    if len(row_numbers) != total_records:
        LOGGER.error("Internal error: row number metadata is inconsistent with records")
        return 1
    if total_records == 0:
        LOGGER.warning("The workbook does not contain any rows to process.")
        return 0

    workers = max(1, int(args.workers))
    LOGGER.info(
        "Processing %s record(s) using %s worker(s). URL column='%s', ID column='%s'",
        total_records,
        workers,
        config.url_column,
        config.id_column,
    )

    results: List[ProcessResult] = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [
            executor.submit(process_record, row_number, record, config)
            for row_number, record in zip(row_numbers, records)
        ]
        for future in tqdm(
            concurrent.futures.as_completed(futures),
            total=len(futures),
            unit="record",
            desc="Processing exhibits",
        ):
            try:
                result = future.result()
            except Exception as error:  # pragma: no cover - defensive branch
                LOGGER.error("Unhandled error: %s", error)
                continue
            results.append(result)

    results.sort(key=lambda item: item.row_number)

    manifest = build_manifest(results)

    manifest_path = args.manifest or (config.out_dir / "manifest.json")
    dump_manifest(manifest, manifest_path)
    LOGGER.info("Manifest written to %s", manifest_path)

    csv_path = args.csv or (config.out_dir / "manifest.csv")
    try:
        write_csv_summary(manifest, csv_path)
        LOGGER.info("CSV summary written to %s", csv_path)
    except Exception as error:  # pragma: no cover - filesystem issues
        LOGGER.warning("Failed to write CSV summary: %s", error)

    summary = manifest["summary"]
    LOGGER.info(
        "Completed: %s success, %s skipped, %s error out of %s rows",
        summary["success"],
        summary["skipped"],
        summary["error"],
        summary["total"],
    )

    return 0 if summary["error"] == 0 else 2


def main() -> None:
    """``setuptools`` entry-point compatible wrapper."""

    raise SystemExit(run())


if __name__ == "__main__":  # pragma: no cover - manual execution
    main()
